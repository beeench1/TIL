from openpyxl import load_workbook

data=[]
cdata=[]
adata_loc='F'
cdata_loc='B'

wb=load_workbook("allData.xlsx",data_only=True)
col=wb['Sheet1'][adata_loc][1:]
for cell in col:
    if cell.value==None:
        break
    else:
        data.append(cell.value)

# 현재 데이터 형성
wb1=load_workbook("data.xlsx",data_only=True)
col=wb1['Sheet1'][cdata_loc]

for cell in col:
    cdata.append(cell.value)

print(cdata[0:5])
print(data[-5])

print(len(cdata))
print(len(data))