from functools import reduce
import random
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import numpy  as np
import pandas as pd
import matplotlib.pyplot as plt

def ch(start,current):
    try:
        ch_=(float(current)-float(start))/float(start)*100
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001


def Storage(data,period,future):
    currentStance='none'
    forward=period
    x=len(data)-period   # 가격 데이터 길이 - 패턴간격
    
    while forward<x:      # 전진값이 커져서 더이상 패턴간격만큼의 패턴을 못만들경우까지
        pattern=[]
        for i in range(period):         
            p=ch(data[forward],data[forward-period+i])
            pattern.append(round(p,2))
        
        patternAr.append(pattern)

        currentPoint=data[forward]
        outcomeRange=data[forward:forward+future]   # 현재 기준 20~30 간격 뒤의 결과값 예측

        try:
            avgOutcome=reduce(lambda x,y : x+y,outcomeRange) /len(outcomeRange)
        except Exception:
            print(str(e))
            avgOutcome=0
        
        futureOutcome=ch(currentPoint,avgOutcome)
        performanceAr.append(round(futureOutcome,2))
        
        forward+=1

def CurrentPattern(data,period):
    for i in range(period):
        c_p=ch(data[-period],data[-(period-i)])
        c_patternAr.append(round(c_p,2))

def patternRecognition(period):
    patFound=0
    for i in range(len(patternAr)):
        sim= np.corrcoef(patternAr[i],c_patternAr)[0,1]
        similar.append(round(sim,2))

        if sim>0.8:
            sim_pattern=patternAr[i]
            plotPatAr.append(sim_pattern)
            plotPatCor.append(round(sim,2))
            plotPatIndex.append(i)
            patFound+=1

    print("찾은 패턴 계수 : " + str(patFound))
    return patFound


# 리스트 저장소 생성
similar=[]
data=[]             # 데이터 
cdata=[]            # 현재 데이터
patternAr=[]        # 패턴 추출 저장
performanceAr=[]    # 패턴의 미래 수익
c_patternAr=[]      # 현시점 패턴
plotPatAr=[]        # 유사한 패턴 저장
plotPatCor=[]          # 유사패턴 상관계수
plotPatIndex=[]     # 유사패턴 인덱스
FoundedOutcome=[]   # 유사 패턴 수익률
FO_Index=[]         # 유사 패턴 수익률 인덱스


code_='gold'
time='60'
cdata_loc='B'
period=120
future=48

code=code_+time
if time=='10': x1=201700 
elif time=='60': x1=33642
elif time=='240': x1=8774
# 데이터 형성
# 행 개수 -> 10m : 201700   ///  60m : 33642  /// 240m : 8774   /// day : 3644
CodeList={"euro10":1,'euro60':2,'euro240':3,'euro_d':4,
            'gold10':5,'gold60':6,'gold240':7,'gold_d':8,
            'oil10':9,'oil60':10,'oil240':11,'oil_d':12}

wb=load_workbook("allData.xlsx",data_only=True)
ws=wb['Sheet1']
for i in range(2,x1):
    data.append(ws.cell(i,CodeList[code]).value)


# 현재 데이터 형성
wb1=load_workbook("data.xlsx",data_only=True)
col=wb1['Sheet1'][cdata_loc]

for cell in col:
    cdata.append(cell.value)

# 패턴 데이터 형성
Storage(data,period,future)
CurrentPattern(cdata,period)

# 패턴인식
patternRecognition(period)

print("데이터 개수 : " + str(len(data)))
print("패턴 리스트의 개수 : "+ str(len(patternAr)))
print("패턴당 데이터 수 :" + str(len(patternAr[0])))
print("현재 패턴의 데이터 수 :" + str(len(c_patternAr)))


# 미래 수익률 계산
for i in plotPatIndex:
    FoundedOutcome.append(performanceAr[i])
    FO_Index.append(period+5)

# 그래프 그리기
plt.figure()
plt.plot(c_patternAr,color='red')
plt.xlabel('period')
plt.ylabel('ch')
plt.scatter(FO_Index,FoundedOutcome,c='#24bc00',alpha=0.4)
plt.show()

# for i in plotPatIndex:
#     FoundedPat=patternAr[i]
#     plt.figure(i)
#     plt.plot(c_patternAr,color='red',label='Current')
#     plt.plot(FoundedPat,color='green',label='Past')
#     plt.xlabel('period')
#     plt.ylabel('ch')
#     plt.text(25,(plt.ylim()[1]+plt.ylim()[0])/3,"COR :" + str(similar[i]))
#     plt.text(25,(plt.ylim()[1]+plt.ylim()[0])*2/3,"Profit(%) : " + str(performanceAr[i]))
#     plt.scatter(FO_Index[i],FoundedOutcome[i],c='#24bc00',alpha=0.4)
#     plt.legend(loc="lower left",ncol=1)
#     plt.show()