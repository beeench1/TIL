import sys
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import *
import time
import pandas as pd
import numpy
import math as m
from datetime import datetime
import csv
import threading
import time
from openpyxl import load_workbook

class Operation(QAxWidget):
    def __init__(self):
        super().__init__()
        self.API_Instance()   
        self.set_Tr()         
        self.set_Tel()             
        
    def API_Instance(self):
        self.setControl("KHOPENAPI.KHOpenAPICtrl.1")

    
    def set_Tr(self):
        self.OnReceiveTrData.connect(self.OnReceiveTrData_)
        
    
    def set_Tel(self):
        self.OnEventConnect.connect(self.OnEventConnect_)

    
    def OnEventConnect_(self,errCode):
        if errCode==0:
            print("connected")
        else:
            print("disconnected")
        
        self.login_event_loop.exit()
    

    def OnReceiveTrData_(self,ScrNo,RQName,TrCode,RecordName,PreNext):
        try:
            if RQName=="week_req":
                self.week={'num':[],"Close":[],"High":[],"Low":[],"Open":[],"Volume":[]}
                for i in range(300):
                    Close=self.GetCommData(TrCode,RQName,i,"현재가")
                    High=self.GetCommData(TrCode,RQName,i,"고가")
                    Low=self.GetCommData(TrCode,RQName,i,"저가")
                    Open=self.GetCommData(TrCode,RQName,i,"시가")
                    volume=self.GetCommData(TrCode,RQName,i,"거래량")

                    self.week['num'].append(i)
                    self.week['Close'].append(abs(float(Close)))
                    self.week['High'].append(abs(float(High)))
                    self.week['Low'].append(abs(float(Low)))
                    self.week['Open'].append(abs(float(Open)))
                    self.week['Volume'].append(abs(int(volume))) 
            
            
            elif RQName=="day_req":
                self.day={'num':[],"Close":[],"High":[],"Low":[],"Open":[],"Volume":[]}
                for i in range(300):
                    Close=self.GetCommData(TrCode,RQName,i,"현재가")
                    High=self.GetCommData(TrCode,RQName,i,"고가")
                    Low=self.GetCommData(TrCode,RQName,i,"저가")
                    Open=self.GetCommData(TrCode,RQName,i,"시가")
                    volume=self.GetCommData(TrCode,RQName,i,"거래량")

                    self.day['num'].append(i)
                    self.day['Close'].append(abs(float(Close)))
                    self.day['High'].append(abs(float(High)))
                    self.day['Low'].append(abs(float(Low)))
                    self.day['Open'].append(abs(float(Open)))
                    self.day['Volume'].append(abs(int(volume)))  


            elif RQName=="min_req":
                self.min={'num':[],"Close":[],"High":[],"Low":[],"Open":[],"Volume":[]}
                for i in range(300):
                    Close=self.GetCommData(TrCode,RQName,i,"현재가")
                    High=self.GetCommData(TrCode,RQName,i,"고가")
                    Low=self.GetCommData(TrCode,RQName,i,"저가")
                    Open=self.GetCommData(TrCode,RQName,i,"시가")
                    volume=self.GetCommData(TrCode,RQName,i,"거래량")

                    self.min['num'].append(i)
                    self.min['Close'].append(abs(float(Close)))
                    self.min['High'].append(abs(float(High)))
                    self.min['Low'].append(abs(float(Low)))
                    self.min['Open'].append(abs(float(Open)))
                    self.min['Volume'].append(abs(int(volume)))  
        except ValueError:
            pass

        self.tr_event_loop.exit()

    
    def CommConnect(self):
        self.dynamicCall("CommConnect()")
        self.login_event_loop=QEventLoop()
        self.login_event_loop.exec_()

    
    def SetInputValue(self,sID,sValue):
        self.dynamicCall("SetInputValue(QString,QString)",sID,sValue)

    
    def CommRqData(self,RQName,TRCode,PrevNext,ScreenNo):
        self.dynamicCall("CommRqData(QString,QString,QString,QString)",
                        RQName,TRCode,PrevNext,ScreenNo)
        self.tr_event_loop=QEventLoop()
        self.tr_event_loop.exec_()

    def MA(self,df,n):
        df['MA'+str(n)]=round(df['Close'].rolling(window=n).mean(),0)

    def HLC(self,df):
        a1=[]
        for i in range(300):
            HLC=(df['High'][i]+df['Low'][i]+df['Close'][i])/3
            a1.append(round(HLC,0))
            
        HLC_=pd.Series(a1)
        df['HLC']=HLC_

    def ATR(self,df, n):
        a1= []
        for i in range(200):
            x=[abs(df['High'][i]-df['Low'][i]),abs(df['High'][i]-df['Close'][i+1]),abs(df['Low'][i]-df['Close'][i+1])]
            TR=max(x)
            a1.append(TR)

        a2 = pd.Series(a1)
        a2_sorted=a2.sort_index(0,ascending=False)
        ATR = a2_sorted.rolling(window=n).mean()
        DE_ATR=ATR.rolling(window=n).mean()
        df['ATR'+str(n)]=ATR
        df['DE_ATR'+str(n)]=DE_ATR


    def Array_Test(self,a1,a2,a3):
        if a1>a2 and a2>a3:
            return int(1)
        elif a1<a2 and a2<a3:
            return int(-1)
        else:
            return int(0)

    def HLC_Test(self,avg,H,L):
        if avg>H:
            return int(1)
        elif avg<L:
            return int(-1)
        else:
            return int(0)
    
    def Vol_Test(self,ATR,ATR2):
        if ATR>ATR2:
            return int(2)
        else:
            return int(1)

    def df_factory(self,data):
        df=pd.DataFrame(data,index=data['num'],
                    columns=['Close','High','Low','Open','Volume'])
        df1=df.sort_index(0,ascending=False)

        self.MA(df1,20)
        self.MA(df1,60)
        self.MA(df1,120)
        self.HLC(df1)
        self.ATR(df1,1)
        self.ATR(df1,20)

        return df1

    def Test(self,Code_Num):
        try:
            Code = str(Code_Num)
            Today=datetime.now().strftime('%Y%m%d')

            # 종목명
            stock_name=kiwoom.GetMasterCodeName(Code)

            # 주봉 데이터
            kiwoom.SetInputValue("종목코드",Code)
            kiwoom.SetInputValue("기준일자",Today)
            kiwoom.SetInputValue("끝일자","0")
            kiwoom.SetInputValue("수정주가구분","0")
            kiwoom.CommRqData("week_req","opt10082",0,"0101")

            data=kiwoom.week
            df_week=kiwoom.df_factory(data)
            
            # 일봉 데이터
            kiwoom.SetInputValue("종목코드",Code)
            kiwoom.SetInputValue("기준일자",Today)
            kiwoom.SetInputValue("수정주가구분","0")
            kiwoom.CommRqData("day_req","opt10081",0,"0102")
            
            data=kiwoom.day
            df_day=kiwoom.df_factory(data)
            
            # 분봉 데이터(10)
            kiwoom.SetInputValue("종목코드",Code)
            kiwoom.SetInputValue("틱범위","10")
            kiwoom.SetInputValue("수정주가구분","0")
            kiwoom.CommRqData("min_req","opt10080",0,"0103")

            data=kiwoom.min
            df_min_10=kiwoom.df_factory(data)
            
            # 분봉 데이터(60)
            kiwoom.SetInputValue("종목코드",Code)
            kiwoom.SetInputValue("틱범위","60")
            kiwoom.SetInputValue("수정주가구분","0")
            kiwoom.CommRqData("min_req","opt10080",0,"0104")

            data=kiwoom.min
            df_min_60=kiwoom.df_factory(data)

            '''
            주봉 : df_week
            일봉 : df_day
            10분봉 : df_min_10
            60분봉 : df_min_60
            Variable Col
            - Close,High,Low,Open
            - Volume
            - MA20, MA60, MA120
            - HLC
            Array_Test(self,a1,a2,a3)
            HLC_Test(self,avg,H,L)
            Vol_Test(self,ATR,ATR2)
            '''

            #Week Array
            week_ma20=df_week['MA20'][0]                     
            week_ma60=df_week['MA60'][0]
            week_ma120=df_week['MA120'][0]

            #Day Array
            day_ma20=df_day['MA20'][0]
            day_ma60=df_day['MA60'][0]
            day_ma120=df_day['MA120'][0]

            #60min Array
            min60_ma20=df_min_60['MA20'][0]
            min60_ma60=df_min_60['MA60'][0]
            min60_ma120=df_min_60['MA120'][0]

            #10min Array
            min10_ma20=df_min_10['MA20'][0]
            min10_ma60=df_min_10['MA60'][0]
            min10_ma120=df_min_10['MA120'][0]

            #Day Vol
            ATR=df_day['ATR1'][0]
            DE_ATR=df_day['DE_ATR20'][0]

            #Day HLC
            H=df_day['High'][2]
            L=df_day['Low'][2]
            C=df_day['Close'][2]
            avg=df_day['HLC'][1]

            Week_Array_Test=kiwoom.Array_Test(week_ma20, week_ma60 ,week_ma120)
            Day_Array_Test=kiwoom.Array_Test(day_ma20, day_ma60, day_ma120)
            min60_Array_Test=kiwoom.Array_Test(min60_ma20, min60_ma60, min60_ma120)
            min10_Array_Test=kiwoom.Array_Test(min10_ma20, min10_ma60, min10_ma120)
            HLC_Test=kiwoom.HLC_Test(avg,H,L)

            Vol_Test=kiwoom.Vol_Test(ATR,DE_ATR)

            score=float(( (Week_Array_Test*0.2) + (Day_Array_Test*0.2) + (min60_Array_Test*0.2) + (min10_Array_Test*0.2) + (HLC_Test)*0.2 )*Vol_Test)
            w_score=float(( (Week_Array_Test*4/13) + (Day_Array_Test*3/13) + (min60_Array_Test*2/13) + (min10_Array_Test*1/13) + (HLC_Test)*3/13 )*Vol_Test)

            score_=round(score,2)
            w_score_=round(w_score,2)

            if score>=1 or w_score>=1:
                self.my_list.append(stock_name)

            '''
            print('종목명 :' + stock_name)
            print('주봉 배열 :' + str(Week_Array_Test))
            print('일봉 배열 :' + str(Day_Array_Test))
            print('60분 배열 :' + str(min60_Array_Test))
            print('10분 배열 :' + str(min10_Array_Test))
            print('일 변동성 :' + str(Vol_Test))
            print('일 HLC :' + str(HLC_Test))
            print("Score : " + str(score_))
            print("W_Score :" + str(w_score_))
            print('==========================================')
            '''
        except ValueError:
            pass

        
    # 관심종목 : column = 5 // 순위종목 : column = 1
    def get_code(self,file,column):
        f=open(str(file),'r')
    
        rdr=csv.reader(f)
        self.code_list=[]
        for i in rdr:
            if(len(self.code_list)<20):
                self.code_list.append(i[column][1:])
            else:
                break
        f.close()

        self.code_list=list(filter(None, self.code_list))[1:]

        return self.code_list


    
    
  
if __name__=="__main__":
    app=QApplication(sys.argv)
    kiwoom=Operation()
    kiwoom.CommConnect()
    
    kiwoom.my_list=[]
    kiwoom.get_code('List.csv',3) 
    code_list=kiwoom.code_list    # code_list : 종목코드
    
    # 종목별 점수 테스트
    for i in code_list:          
        kiwoom.Test(i)
        time.sleep(3)

    week=['A','B','C','D','E','F','G']
    weekday=week[datetime.today().weekday()]
    
    xl_file=load_workbook('MyList.xlsx',data_only=True)
    ws=xl_file['Sheet1']
    #ws.delete_rows(0,50)
    ws[weekday+str(1)]=str(datetime.now().strftime('%Y/%m/%d'))   
    for i in range(len(kiwoom.my_list)):    # my_list : score기반으로 선정한 종목
        ws[weekday+str(i+2)]=kiwoom.my_list[i]

    xl_file.save('MyList.xlsx')
    xl_file.close()

    print("========= Finished =========")
    
