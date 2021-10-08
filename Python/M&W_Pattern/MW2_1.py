import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import statistics as st
import time

W_Pattern=[
[2,1,4,3,5],
[2,1,5,3,4],
[3,1,4,2,5],
[3,1,5,2,4],
[3,2,4,1,5],
[3,2,5,1,4],
[4,1,3,2,5],
[4,1,5,2,3],
[4,2,3,1,5],
[4,2,5,1,3],
[4,3,5,1,2],
[5,1,3,2,4],
[5,1,4,2,3],
[5,2,3,1,4],
[5,2,4,1,3],
[5,3,4,1,2]]

M_Pattern=[
[1,3,2,5,4],
[1,4,2,5,3],
[1,4,3,5,2],
[1,5,2,4,3],
[1,5,3,4,2],
[2,3,1,5,4],
[2,4,1,5,3],
[2,4,3,5,1],
[2,5,1,4,3],
[2,5,3,4,1],
[3,4,1,5,2],
[3,4,2,5,1],
[3,5,1,4,2],
[3,5,2,4,1],
[4,5,1,3,2],
[4,5,2,3,1]]

# Data.xlsx (과거 -> 현재) 내림차순
'''
Data Location
{"euro10":'A','euro60':'B','euro240':'C','euro_d':'D'}
{'gold10':'E','gold60':'F','gold240':'G','gold_d':'H'}
'''

# 데이터 형성
data=[]
wb=load_workbook("Data.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['B'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)

def ch(start,current):
    try:
        ch_=abs(100*(float(current)-float(start))/float(start))
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def ch_(start,current):
    try:
        ch_=100*(float(current)-float(start))/float(start)
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def perPattern(data,patternName,patternNum,period,distance):
    Storage=[]
    Profit=[]
    patternStorage=[]
    patternProfit=[]

    for i in range(0,len(data)-200,120):
        sp=i
        ep=sp+1
        x=[]
        phase=0
        while ep<len(data)-200:
            # Condition of Phase1
            if phase==0 :
                ep=sp+distance
                point=st.mean(data[sp:ep])
                x.append(point)
                #print("Phase1"+"="+str(sp)+":"+str(ep))
                sp=ep+1
                phase+=1
                
            
            # Condition of Phase2
            elif phase==1 :
                ep=sp+distance
                point=st.mean(data[sp:ep])
                x.append(point)
                #print("Phase2"+"="+str(sp)+":"+str(ep))
                sp=ep+1
                phase+=1
            
            # Condition of Phase3
            elif phase==2 :
                ep=sp+distance
                point=st.mean(data[sp:ep])
                x.append(point)
                #print("Phase3"+"="+str(sp)+":"+str(ep))
                sp=ep+1
                phase+=1

            # Condition of Phase4
            elif phase==3 :
                ep=sp+distance
                point=st.mean(data[sp:ep])
                x.append(point)
                #print("Phase4"+"="+str(sp)+":"+str(ep))
                sp=ep+1
                phase+=1


            # Condition of Phase5
            elif phase==4 :
                ep=sp+distance
                point=st.mean(data[sp:ep])
                x.append(point)
                Storage.append(x)
                #print("Phase5"+"="+str(sp)+":"+str(ep))

                currentPoint=data[ep]
                outcomeRange=data[ep:ep+period]   # 현재 기준 20~30 간격 뒤의 결과값 예측

                try:
                    # avgOutcome=reduce(lambda x,y : x+y,outcomeRange) /len(outcomeRange)
                    avgOutcome=st.mean(outcomeRange)
                    
                except Exception:
                    avgOutcome=0
                    pass
                    
                
                futureOutcome=ch_(currentPoint,avgOutcome)
                if abs(futureOutcome)>50:
                    futureOutcome=0
                Profit.append(round(futureOutcome,2))

                #print("Phase5"+"="+str(sp)+":"+str(ep))
                #print("---------------------------------")
                
                sp=ep+1
                phase=0
                
                x=[]

    for i in Storage:
        x=[]
        i_=pd.Series(i)
        x.append(i_.rank()[0])
        x.append(i_.rank()[1])
        x.append(i_.rank()[2])
        x.append(i_.rank()[3])
        x.append(i_.rank()[4])
        patternStorage.append(x)

    # M 패턴별 수익률 저장
    
    
    for j in range(len(patternStorage)):
        x=[]
        if patternName=='W' \
            and W_Pattern[patternNum]==patternStorage[j]: # M패턴과 저장된 패턴 수익률 비교
            patternProfit.append(Profit[j])
        elif patternName=='M' \
            and M_Pattern[patternNum]==patternStorage[j]:
            patternProfit.append(Profit[j])
        
        

    return patternProfit


# Data.xlsx (과거 -> 현재) 내림차순
'''
Data Location
{"euro10":'A','euro60':'B','euro240':'C','euro_d':'D'}
{'gold10':'E','gold60':'F','gold240':'G','gold_d':'H'}
'''

# 데이터 형성
data=[]
wb=load_workbook("Data.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['B'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)

# def perPattern(data,patternName,patternNum,period,distance):

period=[24,48,72,96,120]
distance=[6,12,18,30]
patternName='W'
patternNum=13
start=time.time()

for num in range(0,3):
    print("===================================")
    print("Pattern : " + patternName+str(num))
    for i in distance:
        for j in period:
            a=perPattern(data,'W',num,j,i)
            try:
                result=round(st.mean(a),2)
            except:
                result=0
                continue
            print("인터벌 기준 :" + str(i))
            print("예측 기간 :" + str(j))
            print("수익률 :" + str(result))
            print("===================================")

    end=time.time()
    takeTime=(end-start)/60
    print("Finish : "+str(takeTime)+'min')