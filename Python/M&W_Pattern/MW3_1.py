import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import statistics as st
import time
import math

W_Pattern=[
[2,1,4,3,5],
[2,1,5,3,4],
[3,1,4,2,5],
[3,1,5,2,4],
[3,2,4,1,5],
[3,2,5,1,4],
[4,1,3,2,5],
[4,1,5,2,3],
[4,2,3,1,5],
[4,2,5,1,3],
[4,3,5,1,2],
[5,1,3,2,4],
[5,1,4,2,3],
[5,2,3,1,4],
[5,2,4,1,3],
[5,3,4,1,2]]

M_Pattern=[
[1,3,2,5,4],
[1,4,2,5,3],
[1,4,3,5,2],
[1,5,2,4,3],
[1,5,3,4,2],
[2,3,1,5,4],
[2,4,1,5,3],
[2,4,3,5,1],
[2,5,1,4,3],
[2,5,3,4,1],
[3,4,1,5,2],
[3,4,2,5,1],
[3,5,1,4,2],
[3,5,2,4,1],
[4,5,1,3,2],
[4,5,2,3,1]]

# Data.xlsx (과거 -> 현재) 내림차순
'''
Data Location
{"euro10":'A','euro60':'B','euro240':'C','euro_d':'D'}
{'gold10':'E','gold60':'F','gold240':'G','gold_d':'H'}
'''

# 데이터 형성
data=[]
wb=load_workbook("Data.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['B'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)

def ch(start,current):
    try:
        ch_=abs(100*(float(current)-float(start))/float(start))
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def ch_(start,current):
    try:
        ch_=100*(float(current)-float(start))/float(start)
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def perPattern_M(data,patternName,patternNum,period,cnt):
    Storage=[]
    Profit=[]
    patternStorage=[]
    patternProfit=[]

    for i in range(cnt,len(data)-200,120):
        sp=i
        ep=sp+cnt
        x=[]
        phase=0
        while ep<len(data)-200:
            # Condition of Phase1
            if phase==0 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if min(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt
            
            # Condition of Phase2
            elif phase==1 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if max(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt
            
            # Condition of Phase3
            elif phase==2 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if min(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt

            # Condition of Phase4
            elif phase==3 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if max(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt


            # Condition of Phase5
            elif phase==4 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if min(sample)==data[m]:
                    x.append(point)
                    Storage.append(x)

                    currentPoint=data[ep]
                    outcomeRange=data[ep:ep+period]   # 현재 기준 20~30 간격 뒤의 결과값 예측

                    try:
                        # avgOutcome=reduce(lambda x,y : x+y,outcomeRange) /len(outcomeRange)
                        avgOutcome=st.mean(outcomeRange)
                        
                    except Exception:
                        avgOutcome=0
                        pass
                        
                    
                    futureOutcome=ch_(currentPoint,avgOutcome)
                    if abs(futureOutcome)>50:
                        futureOutcome=0
                    Profit.append(round(futureOutcome,2))

                    #print("Phase5"+"="+str(sp)+":"+str(ep))
                    #print("---------------------------------")
                    phase=0
                    sp=ep+1
                    ep=sp+cnt
                
                    x=[]
                else:
                    sp=sp+1
                    ep=sp+cnt

    for i in Storage:
        x=[]
        i_=pd.Series(i)
        x.append(i_.rank()[0])
        x.append(i_.rank()[1])
        x.append(i_.rank()[2])
        x.append(i_.rank()[3])
        x.append(i_.rank()[4])
        patternStorage.append(x)

    # M 패턴별 수익률 저장
    
    
    for j in range(len(patternStorage)):
        x=[]
        if patternName=='W' \
            and W_Pattern[patternNum]==patternStorage[j]: # M패턴과 저장된 패턴 수익률 비교
            patternProfit.append(Profit[j])
        elif patternName=='M' \
            and M_Pattern[patternNum]==patternStorage[j]:
            patternProfit.append(Profit[j])
        
        

    return patternProfit, len(patternStorage), len(patternProfit)

def perPattern_W(data,patternName,patternNum,period,cnt):
    Storage=[]
    Profit=[]
    patternStorage=[]
    patternProfit=[]

    for i in range(cnt,len(data)-200,120):
        sp=i
        ep=sp+cnt
        x=[]
        phase=0
        while ep<len(data)-200:
            # Condition of Phase1
            if phase==0 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if max(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt
            
            # Condition of Phase2
            elif phase==1 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if min(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt
            
            # Condition of Phase3
            elif phase==2 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if max(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt

            # Condition of Phase4
            elif phase==3 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if min(sample)==data[m]:
                    x.append(point)
                    phase+=1
                    sp=ep+1
                    ep=sp+cnt
                else:
                    sp=sp+1
                    ep=sp+cnt


            # Condition of Phase5
            elif phase==4 :
                sample=data[sp:ep]
                m=math.ceil((sp+ep)/2)
                point = data[m]
                if max(sample)==data[m]:
                    x.append(point)
                    Storage.append(x)

                    currentPoint=data[ep]
                    outcomeRange=data[ep:ep+period]   # 현재 기준 20~30 간격 뒤의 결과값 예측

                    try:
                        # avgOutcome=reduce(lambda x,y : x+y,outcomeRange) /len(outcomeRange)
                        avgOutcome=st.mean(outcomeRange)
                        
                    except Exception:
                        avgOutcome=0
                        pass
                        
                    
                    futureOutcome=ch_(currentPoint,avgOutcome)
                    if abs(futureOutcome)>50:
                        futureOutcome=0
                    Profit.append(round(futureOutcome,2))

                    #print("Phase5"+"="+str(sp)+":"+str(ep))
                    #print("---------------------------------")
                    phase=0
                    sp=ep+1
                    ep=sp+cnt

                    x=[]
                else:
                    sp=sp+1
                    ep=sp+cnt

    for i in Storage:
        x=[]
        i_=pd.Series(i)
        x.append(i_.rank()[0])
        x.append(i_.rank()[1])
        x.append(i_.rank()[2])
        x.append(i_.rank()[3])
        x.append(i_.rank()[4])
        patternStorage.append(x)

    # M 패턴별 수익률 저장
    
    
    for j in range(len(patternStorage)):
        x=[]
        if patternName=='W' \
            and W_Pattern[patternNum]==patternStorage[j]: # M패턴과 저장된 패턴 수익률 비교
            patternProfit.append(Profit[j])
        elif patternName=='M' \
            and M_Pattern[patternNum]==patternStorage[j]:
            patternProfit.append(Profit[j])
        
        

    return patternProfit, len(patternStorage), len(patternProfit)



# Data.xlsx (과거 -> 현재) 내림차순
'''
Data Location
{"euro10":'A','euro60':'B','euro240':'C','euro_d':'D'}
{'gold10':'E','gold60':'F','gold240':'G','gold_d':'H'}
'''

# 데이터 형성
data=[]
wb=load_workbook("Data.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['B'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)

# def perPattern(data,patternName,patternNum,period,distance):

period=[24,48,72,96,120]    # 예측기간
distance=[5,13,23,47]   # 캔들 개수
patternName='M'
patternNum=13
start=time.time()
print("start")

for num in range(9,10):
    print("===================================")
    print("Pattern : " + patternName+str(num))
    for i in distance:
        for j in period:
            a=perPattern_M(data,'M',num,j,i)
            try:
                result=round(st.mean(a[0]),2)
            except:
                result=0
                
            print("인터벌 기준 :" + str(i))
            print("예측 기간 :" + str(j))
            print("수익률 :" + str(result))
            print("표본 수 : " + str(a[1]) + " 맞는 패턴 개수 :" + str(a[2]))
            print("===================================")

    end=time.time()
    takeTime=(end-start)/60
    print("Finish : "+str(takeTime)+'min')