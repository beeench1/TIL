import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import statistics as st

W_Pattern=[
[2,1,4,3,5],
[2,1,5,3,4],
[3,1,4,2,5],
[3,1,5,2,4],
[3,2,4,1,5],
[3,2,5,1,4],
[4,1,3,2,5],
[4,1,5,2,3],
[4,2,3,1,5],
[4,2,5,1,3],
[4,3,5,1,2],
[5,1,3,2,4],
[5,1,4,2,3],
[5,2,3,1,4],
[5,2,4,1,3],
[5,3,4,1,2]]

M_Pattern=[
[1,3,2,5,4],
[1,4,2,5,3],
[1,4,3,5,2],
[1,5,2,4,3],
[1,5,3,4,2],
[2,3,1,5,4],
[2,4,1,5,3],
[2,4,3,5,1],
[2,5,1,4,3],
[2,5,3,4,1],
[3,4,1,5,2],
[3,4,2,5,1],
[3,5,1,4,2],
[3,5,2,4,1],
[4,5,1,3,2],
[4,5,2,3,1]]

# Data.xlsx (과거 -> 현재) 내림차순
'''
Data Location
{"euro10":'A','euro60':'B','euro240':'C','euro_d':'D'}
{'gold10':'E','gold60':'F','gold240':'G','gold_d':'H'}
'''

# 데이터 형성
data=[]
wb=load_workbook("Data.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['B'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)

        
def ch(start,current):
    try:
        ch_=abs(100*(float(current)-float(start))/float(start))
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def ch_(start,current):
    try:
        ch_=100*(float(current)-float(start))/float(start)
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

patternName='W'
refVol=0.1  # 전환비율
period=24   # 예측 기간
distance=5  # 패턴 간격
phase=0
Storage=[]  # 패턴 리스트
Profit=[]   # 패턴 수익률 저장


# 패턴 만들기
for i in range(0,len(data)-200,120):
    sp=i
    ep=sp+1
    x=[]
    while ep<len(data)-200:
        # Condition of Phase1
        if phase==0 :
            ep=sp+distance
            point=st.mean(data[sp:ep])
            x.append(point)
            print("Phase1"+"="+str(sp)+":"+str(ep))
            sp=ep+1
            phase+=1
            
        
        # Condition of Phase2
        elif phase==1 :
            ep=sp+distance
            point=st.mean(data[sp:ep])
            x.append(point)
            print("Phase2"+"="+str(sp)+":"+str(ep))
            sp=ep+1
            phase+=1
        
        # Condition of Phase3
        elif phase==2 :
            ep=sp+distance
            point=st.mean(data[sp:ep])
            x.append(point)
            print("Phase3"+"="+str(sp)+":"+str(ep))
            sp=ep+1
            phase+=1

        # Condition of Phase4
        elif phase==3 :
            ep=sp+distance
            point=st.mean(data[sp:ep])
            x.append(point)
            print("Phase4"+"="+str(sp)+":"+str(ep))
            sp=ep+1
            phase+=1


        # Condition of Phase5
        elif phase==4 :
            ep=sp+distance
            point=st.mean(data[sp:ep])
            x.append(point)
            Storage.append(x)
            print("Phase5"+"="+str(sp)+":"+str(ep))

            currentPoint=data[ep]
            outcomeRange=data[ep:ep+period]   # 현재 기준 20~30 간격 뒤의 결과값 예측

            try:
                # avgOutcome=reduce(lambda x,y : x+y,outcomeRange) /len(outcomeRange)
                avgOutcome=st.mean(outcomeRange)
                
            except Exception:
                avgOutcome=0
                pass
                
            
            futureOutcome=ch_(currentPoint,avgOutcome)
            if abs(futureOutcome)>50:
                futureOutcome=0
            Profit.append(round(futureOutcome,2))

            print("Phase5"+"="+str(sp)+":"+str(ep))
            print("---------------------------------")
            
            sp=ep+1
            phase=0
            
            x=[]
        
print(len(Storage))
print(len(Profit))
patternStorage=[]   # 일반화 패턴 리스트
patternProfit=[]    # 패턴 수익률 리스트

# 전환점 랭크를 통한 패턴 일반화
for i in Storage:
    x=[]
    i_=pd.Series(i)
    x.append(i_.rank()[0])
    x.append(i_.rank()[1])
    x.append(i_.rank()[2])
    x.append(i_.rank()[3])
    x.append(i_.rank()[4])
    patternStorage.append(x)

# M 패턴별 수익률 저장
for i in range(16):
    x=[]
    for j in range(len(patternStorage)):
        if patternName=='W' \
            and W_Pattern[i]==patternStorage[j]: # M패턴과 저장된 패턴 수익률 비교
            x.append(Profit[j])
        elif patternName=='M' \
            and M_Pattern[i]==patternStorage[j]:
            x.append(Profit[j])
    
    patternProfit.append(x)

print(len(patternProfit))
print(len(patternProfit[0]))
print(len(patternProfit[1]))
print(len(patternProfit[2]))
print(len(patternProfit[3]))
print(len(patternProfit[4]))
print(len(patternProfit[5]))
print(len(patternProfit[6]))
print(len(patternProfit[7]))
print(len(patternProfit[8]))
print(len(patternProfit[9]))
print(len(patternProfit[10]))
print(len(patternProfit[11]))
print(len(patternProfit[12]))
print(len(patternProfit[13]))
print(len(patternProfit[14]))
print(len(patternProfit[15]))

# 패턴별 수익률 시각화
for i in range(len(patternProfit[i])):
    indexS=[]
    for j in range(len(patternProfit[i])): 
        indexS.append(7)

    fig = plt.figure()
    plt.title(str(i))
    ax1 = fig.add_subplot()
    ax2=ax1.twinx()             # Y축 추가
    if patternName=='M':a=M_Pattern[i]
    elif patternName=='W':a=W_Pattern[i]
    ax1.plot(a, marker='s')
    ax2.scatter(indexS,patternProfit[i],c='#24bc00',alpha=0.4)
    ax2.axhline(0,color='red')
    ax2.axhline(st.median(patternProfit[i]),color='green')
    plt.show()

