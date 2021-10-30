import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import statistics as st
import time
import math

def ch_(start,current):
    try:
        ch_=100*(float(current)-float(start))/float(start)
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def permute(arr):
    result = [arr[:]]
    c = [0] * len(arr)
    i = 0
    while i < len(arr):
        if c[i] < i:
            if i % 2 == 0:
                arr[0], arr[i] = arr[i], arr[0]
            else:
                arr[c[i]], arr[i] = arr[i], arr[c[i]]
            result.append(arr[:])
            c[i] += 1
            i = 0
        else:
            c[i] = 0
            i += 1
    return result


def pattern(data,pat,forward,distance,ref):
    point=[]
    winStorage=[]
    loseStorage=[]
    entry=0
    position=False
    sp=forward
    ep=sp+distance
    m=math.ceil(sp+ep)
    
    for i in range(len(data)-200):
        try:
            m=math.ceil((sp+ep)/2)
            cnt=0
            if max(data[sp:ep])==data[m] or min(data[sp:ep])==data[m]:
                
                point.append(data[m])
                if len(point)==6:
                    point.pop(0)
                
            if len(point)==5 and position==False:
                pattern=[]
                series=pd.Series(point)
                pattern.append(series.rank()[0])
                pattern.append(series.rank()[1])
                pattern.append(series.rank()[2])
                pattern.append(series.rank()[3])
                pattern.append(series.rank()[4])
                print(pattern)
                if pat==pattern:
                    entry=data[ep]
                    position=True
                    cnt+=1
                    
            elif position==True:
                cnt+=1
                profit=ch_(entry,data[ep])
                if cnt>ref:
                    if profit>0:
                        winStorage.append(profit)
                        position=False
                    else:
                        loseStorage.append(profit)
                        position=False
                    
            sp+=1
            ep=sp+distance
        except:
            print('error')

    try:
    
        profit=round(st.mean(winStorage),2)
        loss=round(st.mean(loseStorage),2)
        pr=round(len(winStorage)/(len(winStorage)+len(loseStorage)),2)


        print("========= MW Long Simulation =========")
        print(str(pat[0])+str(pat[1])+str(pat[2])+str(pat[3])+str(pat[4]))
        print('수익(%) : '+ str(profit))
        print('손실(%) : '+ str(loss)) 
        print('승률(%) : '+ str(pr))
        print("총 거래수 :" + str(len(winStorage)+len(loseStorage)))
    except:
        print("========= MW Long Simulation =========")
        print(str(pat[0])+str(pat[1])+str(pat[2])+str(pat[3])+str(pat[4]))
        print("NOT FOUNDED")

# 데이터 형성
data=[]
wb=load_workbook("gold60.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['E'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)

print(len(data))

allData=permute([1,2,3,4,5])

# def pattern(data,pat,forward,distance,ref):
pat=[3,2,5,1,4]
start=time.time()
cnt=0
ref=[6,12,18,24,36,48]

pattern(data,pat,100,25,12)
# for i in ref:
#     print(cnt)
#     pattern(data,pat,100,25,i)
#     cnt+=1

end=time.time()
takeTime=round((end-start)/60,0)
print("Finish : "+str(takeTime)+'min')