import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import statistics as st
import time
import math

def ch_(start,current):
    try:
        ch_=100*(float(current)-float(start))/float(start)
        if ch_==0:
            return 0.00000001
        else:
            return ch_
    
    except:
        return 0.00000001

def permute(arr):
    result = [arr[:]]
    c = [0] * len(arr)
    i = 0
    while i < len(arr):
        if c[i] < i:
            if i % 2 == 0:
                arr[0], arr[i] = arr[i], arr[0]
            else:
                arr[c[i]], arr[i] = arr[i], arr[c[i]]
            result.append(arr[:])
            c[i] += 1
            i = 0
        else:
            c[i] = 0
            i += 1
    return result


def pattern_L(data,pat,forward,distance,ref,ref_c):
    point=[]
    winStorage=[]
    loseStorage=[]
    entry=0
    position=False
    sp=forward
    ep=forward+distance
    m=math.ceil(sp+ep)
    
    for i in range(len(data)-200):
        try:
            m=math.ceil((sp+ep)/2)
            if max(data[sp:ep])==data[m] or min(data[sp:ep])==data[m]:
                point.append(data[m])
                if len(point)==6:
                    del point[0]

            if len(point)==5 and position==False:
                pattern=[]
                series=pd.Series(point)
                pattern.append(series.rank()[0])
                pattern.append(series.rank()[1])
                pattern.append(series.rank()[2])
                pattern.append(series.rank()[3])
                pattern.append(series.rank()[4])

                if pat==pattern:
                    entry=data[ep]
                    position=True
                    sp+=distance
                    ep=sp+distance
                    
            elif position==True:
                profit=ch_(entry,data[sp])
                if profit>ref:
                    winStorage.append(profit)
                    position=False
                elif profit<ref_c:
                    loseStorage.append(profit)
                    position=False
                    
            
            sp+=1
            ep=sp+distance
        except:
            pass

    print(len(winStorage))
    print(len(loseStorage))
    
    try:
        profit=round(st.mean(winStorage),2)
        loss=round(st.mean(loseStorage),2)
        pr=round(len(winStorage)/(len(winStorage)+len(loseStorage)),2)
        pl=abs(round(profit/loss,2))
    except:
        profit,loss,pr,pl='None','None','None',0

    return profit,loss,pr,pl

    # if pl>1.2:
    #     with open("result.txt", "a") as f:
    #         f.write("\n")
    #         f.write("========= MW Long Simulation =========\n")
    #         f.write(str(pat[0])+str(pat[1])+str(pat[2])+str(pat[3])+str(pat[4])+"\n")
    #         f.write('수익(%) : '+ str(profit)+"\n")
    #         f.write('손실(%) : '+ str(loss)+"\n")
    #         f.write('손익비 : ' + str(pl)+"\n")
    #         f.write('승률(%) : '+ str(pr)+"\n")
    #         f.write("총 거래수 :" + str(len(winStorage)+len(loseStorage))+"\n")
    #         f.write("================ Finish ================")
            
    #     print("========= MW Long Simulation =========")
    #     print(str(pat[0])+str(pat[1])+str(pat[2])+str(pat[3])+str(pat[4]))
    #     print('수익(%) : '+ str(profit))
    #     print('손실(%) : '+ str(loss))
    #     print('손익비 : ' + str(pl))
    #     print('승률(%) : '+ str(pr))
    #     print("총 거래수 :" + str(len(winStorage)+len(loseStorage)))
    #     print("================ Finish ================")
    # else:
    #     print("--------ing---------")

def pattern_S(data,pat,forward,distance,ref,ref_c):
    point=[]
    winStorage=[]
    loseStorage=[]
    entry=0
    position=False
    sp=forward
    ep=forward+distance
    m=math.ceil(sp+ep)
    
    for i in range(len(data)-200):
        try:
            m=math.ceil((sp+ep)/2)
            if max(data[sp:ep])==data[m] or min(data[sp:ep])==data[m]:
                point.append(data[m])
                if len(point)==6:
                    del point[0]

            if len(point)==5 and position==False:
                pattern=[]
                series=pd.Series(point)
                pattern.append(series.rank()[0])
                pattern.append(series.rank()[1])
                pattern.append(series.rank()[2])
                pattern.append(series.rank()[3])
                pattern.append(series.rank()[4])

                if pat==pattern:
                    entry=data[ep]
                    position=True
                    sp+=distance
                    ep=sp+distance
                    
            elif position==True:
                profit=ch_(data[sp],entry)
                if profit>ref:
                    winStorage.append(profit)
                    position=False
                elif profit<ref_c:
                    loseStorage.append(profit)
                    position=False
                    
            
            sp+=1
            ep=sp+distance
        except:
            pass

    print(len(winStorage))
    print(len(loseStorage))
    
    try:
        profit=round(st.mean(winStorage),2)
        loss=round(st.mean(loseStorage),2)
        pr=round(len(winStorage)/(len(winStorage)+len(loseStorage)),2)
        pl=abs(round(profit/loss,2))
    except:
        profit,loss,pr,pl='None','None','None',0

    return profit, loss, pr, pl
    
    # if pl>1.2:
    #     with open("result.txt", "a") as f:
    #         f.write("\n")
    #         f.write("========= MW Long Simulation =========\n")
    #         f.write(str(pat[0])+str(pat[1])+str(pat[2])+str(pat[3])+str(pat[4])+"\n")
    #         f.write('수익(%) : '+ str(profit)+"\n")
    #         f.write('손실(%) : '+ str(loss)+"\n")
    #         f.write('손익비 : ' + str(pl)+"\n")
    #         f.write('승률(%) : '+ str(pr)+"\n")
    #         f.write("총 거래수 :" + str(len(winStorage)+len(loseStorage))+"\n")
    #         f.write("================ Finish ================")
            
    #     print("========= MW Long Simulation =========")
    #     print(str(pat[0])+str(pat[1])+str(pat[2])+str(pat[3])+str(pat[4]))
    #     print('수익(%) : '+ str(profit))
    #     print('손실(%) : '+ str(loss))
    #     print('손익비 : ' + str(pl))
    #     print('승률(%) : '+ str(pr))
    #     print("총 거래수 :" + str(len(winStorage)+len(loseStorage)))
    #     print("================ Finish ================")
    # else:
    #     print("--------ing---------")



# 데이터 형성
data=[]
wb=load_workbook("gold60.xlsx",data_only=True)    # load_workbook("fileName")
col=wb['Sheet1']['E'][1:]             # object[sheetName][ColName][Range] : 열 데이터 인덱싱
for cell in col:
    if cell.value==None:        # 셀값이 None이면 중단.
        break
    else:
        data.append(cell.value)



allData=permute([1,2,3,4,5])

# def pattern_L(data,pat,forward,distance,ref,ref_c):

start=time.time()

for forward in forward_:
    for distance in distance_:
        for ref in ref_:
            for ratio in ratio_:
                a=pattern_L(data,pat,forward,distance,ref,ref/ratio)



end=time.time()
taketime=(end-start)/60
print(str(taketime) + 'min')