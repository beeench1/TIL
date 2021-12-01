import csv
from openpyxl import load_workbook

f=open('1234.csv','r')
    
rdr=csv.reader(f)
code_list=[]
for i in rdr:
    if(len(code_list)<100):
        code_list.append(i[3][1:])
    else:
        break
f.close()

code_list1=list(filter(None, code_list))[1:]

xl_file=load_workbook('code.xlsx',data_only=True)
ws=xl_file['Sheet1']
for i in range(len(code_list1)):
    ws['A'+str(i+1)]=code_list1[i]

xl_file.save('code.xlsx')
xl_file.close()

print(code_list1)
print(code_list1[2])
print(type(code_list1[2]))
print(len(code_list1))